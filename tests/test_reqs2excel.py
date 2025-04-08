import os
import json
import uuid
import shutil
import tempfile
from pathlib import Path

import pytest
import openpyxl
from unittest.mock import patch

from autoreq.reqs2excel import load_requirements_from_gateway, cli


REQUIREMENTS_EXAMPLE = {
    "0": {
        "id0": {
            "title": "Sample Requirement",
            "description": "This is a sample requirement description.",
            "id": "id0",
            "last_modified": "tbd",
        },
        "id1": {
            "title": "Another Requirement",
            "description": "This is another requirement description.",
            "id": "id1",
            "last_modified": "tbd",
        },
    },
    "1": {
        "id2": {
            "title": "Third Requirement",
            "description": "This is the third requirement description.",
            "id": "id2",
            "last_modified": "tbd",
        },
        "id3": {
            "title": "Fourth Requirement",
            "description": "This is the fourth requirement description.",
            "id": "id3",
            "last_modified": "tbd",
        },
    },
}

MOCK_REQ2FUN_MAPPING = {
    "The system will support 6 tables.": "unknown",
    "The system will support 4 seats per table.": "unknown",
    "The system will support the following entrees:\n\nsteak\nchicken\nlobster\npasta\nnone": "none",
    "Placing an order updates the table's occupied status to true within the table database.": "None",
    "Placing an order updates the table's number in party within the table database.": "None",
    "Placing an order updates the seat's order within the table database.": "None",
    "Placing an order increases the table's check total within the table database, by an amount depending on the entree ordered, according to the following schedule:\n\nEntree: Amount\n\nsteak: 14.0\nchicken: 10.0\nlobster: 18.0\npasta: 12.0\nnone: 0.0": "None",
    "Clearing a table updates the table's occupied status to false within the table database.": "unknown",
    "Clearing a table updates the table's number in party to 0 within the table database.": "None",
    "Clearing a table clears the orders for all seats of the table within the table database.": "None",
    "Clearing a table updates the table's check total to 0.0 within the table database.": "None",
    "The system will provide a way to obtain the check total for a given table.": "unknown",
    "The system will support a waiting list of up to 10 parties.": "None",
    "The system will provide a means of adding a party to the waiting list, with the party specified by name.": "None",
    "The system will provide a means of obtaining the name of the party at the head of the waiting list.": "unknown",
    "Placing certain orders will qualify the seat for free dessert, according to the following schedule:\n\nSteak with caesar salad and a mixed drink qualifies a seat for pie.\nLobster with green salad and wine qualifies a seat for cake.": "None"
}


def _copy_folder(src, dst):
    for item in src.iterdir():
        target = dst / item.name
        if item.is_dir():
            shutil.copytree(item, target)
        else:
            shutil.copy2(item, target)


def check_excel_file_integrity(file_path: Path):
    assert file_path.exists(), "The Excel file should be created."
    assert file_path.suffix == ".xlsx", "The output file should be an Excel file."
    wb = openpyxl.load_workbook(file_path)
    assert set(wb.sheetnames) == {
        "Options",
        "Requirements",
    }, "Workbook should contain 'Options' and 'Requirements' sheets."
    assert wb.active.title == "Requirements", "Active sheet should be 'Requirements'."
    assert (
        wb.worksheets[0].title == "Requirements"
        and wb.worksheets[0].sheet_state == "visible"
    ), "The first sheet should be 'Requirements' and visible."
    assert (
        wb.worksheets[1].title == "Options" and wb.worksheets[1].sheet_state == "hidden"
    ), "The second sheet should be 'Options' and hidden."

    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    reqs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        reqs.append(dict(zip(headers, row)))
    assert all(
        req.get(key) is not None
        for req in reqs
        for key in ("Key", "ID", "Title", "Description", "Module", "Function")
    ), "All requirements should have Key, ID, Title, Description, Module, and Function fields."


def check_excel_file_contents(
    file_path: Path, requirements: list[dict]
):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Requirements"]
    headers = [cell.value for cell in sheet[1]]
    sheet_reqs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sheet_reqs.append(dict(zip(headers, row)))

    for r1, r2 in zip(sheet_reqs, requirements):
        assert (
            r1["Key"] == r2["id"]
        ), f"Expected {r2['id']} but got {r1['Key']} in requirements sheet."
        assert (
            r1["ID"] == r2["id"]
        ), f"Expected {r2['id']} but got {r1['ID']} in requirements sheet."
        assert (
            r1["Title"] == r2["title"]
        ), f"Expected {r2['title']} but got {r1['Title']} in requirements sheet."
        assert (
            r1["Description"] == r2["description"]
        ), f"Expected {r2['description']} but got {r1['Description']} in requirements sheet."


@pytest.fixture
def envs_dir():
    return Path(__file__).parent / "envs_for_tests"


@pytest.fixture
def fake_requirements_dir(tmp_path):
    reqs_dir = tmp_path / "requirements_gateway"
    reqs_dir.mkdir()
    with open(reqs_dir / "requirements.json", "w") as f:
        json.dump(REQUIREMENTS_EXAMPLE, f)
    return reqs_dir


@pytest.fixture
def real_requirements_dir(envs_dir):
    return envs_dir / "requirements_gateway"


@pytest.fixture
def vectorcast_dir():
    _vectorcast_dir = os.getenv("VECTORCAST_DIR")
    assert _vectorcast_dir
    _vectorcast_dir = Path(_vectorcast_dir)
    assert (
        _vectorcast_dir.is_dir()
    ), "VectorCast directory should be set in VECTORCAST_DIR environment variable."
    return _vectorcast_dir


def test_load_requirements_from_gateway(fake_requirements_dir):
    reqs = load_requirements_from_gateway(fake_requirements_dir)
    assert (
        isinstance(reqs, list) and len(reqs) == 4
    ), "load_requirements_from_gateway should return a list of requirements."
    for req in reqs:
        assert isinstance(req, dict) and all(
            key in req for key in ("id", "title", "description")
        ), "Each requirement should have id, title, and description."


def test_main_pre_existing_requirements(monkeypatch, envs_dir, real_requirements_dir):
    current_workdir = os.getcwd()
    env = next(
        (Path(l.strip()) for l in Path(envs_dir, "batch.txt").read_text().splitlines()
         if l.strip())
    )   # no need to test for all envs

    requirements: dict = json.loads((real_requirements_dir / "requirements.json").read_text())
    requirements = next({k: v} for k, v in requirements.items())

    with tempfile.TemporaryDirectory() as tmp_folder:
        out_folder = Path(tmp_folder)
        os.chdir(out_folder)
        _copy_folder(envs_dir, out_folder)

        reqs_dir = out_folder / str(uuid.uuid4())
        reqs_dir.mkdir()
        with open(reqs_dir / "requirements.json", "w") as f:
            json.dump(requirements, f)

        test_args = [
            "reqs2excel",
            str(env.absolute()),
            "--requirements-gateway-path",
            str(reqs_dir.absolute()),
            "--output-file",
            f"output.xlsx",
        ]
        monkeypatch.setattr("sys.argv", test_args)

        with patch("autoreq.trace_reqs2code.Reqs2CodeMapper.map_reqs_to_code_for_env",
                   return_value=MOCK_REQ2FUN_MAPPING):
            cli()

        out_file = out_folder / "output.xlsx"
        check_excel_file_integrity(out_file)
        check_excel_file_contents(
            out_file,
            load_requirements_from_gateway(reqs_dir),
        )

    os.chdir(current_workdir)


def test_main_with_rgw_init(monkeypatch, envs_dir, vectorcast_dir):
    current_workdir = os.getcwd()
    with tempfile.TemporaryDirectory() as out_folder:
        os.chdir(out_folder)
        _copy_folder(envs_dir, Path(out_folder))
        test_args = [
            "reqs2excel",
            "@batch.txt",
            "--init-requirements",
            "--csv-template",
            f"{vectorcast_dir}/examples/RequirementsGW/CSV_Requirements_For_Tutorial.csv",
            "--output-file",
            f"output.xlsx",
        ]
        monkeypatch.setattr("sys.argv", test_args)

        with patch("autoreq.trace_reqs2code.Reqs2CodeMapper.map_reqs_to_code_for_env",
                   return_value=MOCK_REQ2FUN_MAPPING):
            cli()

        out_file = Path(out_folder) / "output.xlsx"
        check_excel_file_integrity(out_file)

        env_dirs = [
            Path(l.strip()).parent
            for l in Path("batch.txt").read_text().splitlines()
            if l.strip()
        ]
        requirements = []
        for ed in env_dirs:
            rgw_path = ed / "requirements_gateway"
            requirements += load_requirements_from_gateway(rgw_path)

        check_excel_file_contents(out_file, requirements)

    os.chdir(current_workdir)

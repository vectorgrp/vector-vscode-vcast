import argparse
import json
from tqdm.asyncio import tqdm_asyncio
from pathlib import Path
import csv
import os
import subprocess
import tempfile
import asyncio
import logging
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from autoreq.requirements_manager import RequirementsManager
from autoreq.util import ensure_env
from autoreq.test_generation.vcast_context_builder import VcastContextBuilder

from .test_generation.environment import Environment
from .requirement_generation.generation import RequirementsGenerator
from .requirement_generation.high_level_generation import HighLevelRequirementsGenerator

_ORDERED_FIELDNAMES = [
    'Key',
    'ID',
    'Module',
    'Title',
    'Description',
    'Function',
    'Lines',
]


def save_requirements_to_csv(requirements, output_file):
    field_names = [
        name
        for name in _ORDERED_FIELDNAMES
        if name in (requirements[0] if requirements else {})
    ]
    with open(output_file, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=field_names)
        writer.writeheader()
        for req in requirements:
            writer.writerow(req)


def fix_worksheet_cols_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max(
                    (dims.get(cell.column, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value


def save_requirements_to_excel(requirements, source_envs, output_file):
    funcs = {}
    for env in source_envs:
        for f in env.testable_functions:
            key = f['unit_name']
            funcs.setdefault(key, [])
            funcs[key].append(f['name'])

    workbook = Workbook()
    main_ws = workbook.active
    main_ws.title = 'Requirements'
    lists_ws = workbook.create_sheet('Options')
    lists_ws.sheet_state = 'hidden'

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='C6EFCE')
    headers = list(requirements[0].keys())
    for col, header in enumerate(headers, start=1):
        cell = main_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, req in enumerate(requirements, start=2):
        for col_idx, header in enumerate(headers, start=1):
            main_ws.cell(row=row_idx, column=col_idx, value=req[header])

    # In column A, list all modules.
    modules = list(funcs.keys())
    for row, module in enumerate(modules, start=1):
        lists_ws.cell(row=row, column=1, value=module)

    # For each module, write its functions in its own column (starting at column B)
    # and create a named range using create_named_range().
    for i, module in enumerate(modules):
        functions = funcs[module] + ['None']
        col = i + 2
        for row, func in enumerate(functions, start=1):
            lists_ws.cell(row=row, column=col, value=func)
        col_letter = get_column_letter(col)
        end_row = len(functions)
        range_ref = f'${col_letter}$1:${col_letter}${end_row}'
        range_name = module.replace(' ', '_').replace('-', '_')
        workbook.create_named_range(range_name, lists_ws, range_ref)

    num_data_rows = len(requirements)
    module_col_index = headers.index('Module') + 1
    function_col_index = headers.index('Function') + 1
    module_col_letter = get_column_letter(module_col_index)
    dv_module = DataValidation(
        type='list', formula1=f'=Options!$A$1:$A${len(modules)}', allow_blank=True
    )
    main_ws.add_data_validation(dv_module)
    dv_module_range = f'{get_column_letter(module_col_index)}2:{get_column_letter(module_col_index)}{num_data_rows + 1}'
    dv_module.add(dv_module_range)

    # Use INDIRECT with SUBSTITUTE to reference the named range based on the Module cell.
    for row in range(2, num_data_rows + 2):
        module_cell = f'{module_col_letter}{row}'
        formula = f'=INDIRECT(SUBSTITUTE({module_cell}," ","_"))'
        dv_func = DataValidation(type='list', formula1=formula, allow_blank=True)
        main_ws.add_data_validation(dv_func)
        function_cell = f'{get_column_letter(function_col_index)}{row}'
        dv_func.add(function_cell)

    fix_worksheet_cols_width(main_ws)
    workbook.save(output_file)


def save_requirements_to_html(requirements, output_file):
    html_content = """
    <html>
    <head>
        <title>Requirements</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; background-color: #ffffff; color: #000000; }
            h1 { color: #2c3e50; }
            h2 { color: #34495e; margin-top: 30px; }
            .requirement { background-color: #f7f7f7; padding: 15px; margin: 10px 0; border-radius: 5px; }
            .req-key { font-weight: bold; color: #2980b9; }
            .req-description { margin-top: 10px; color: #333333; }
        </style>
    </head>
    <body>
        <h1>Requirements</h1>
    """
    # Group requirements by function
    requirements_by_function = {}
    for req in requirements:
        func_name = req['Function']
        requirements_by_function.setdefault(func_name, []).append(req)
    # Generate HTML content
    for func_name, reqs in requirements_by_function.items():
        html_content += f'<h2>{func_name}</h2>'
        for req in reqs:
            html_content += f"""
            <div class="requirement">
                <div class="req-key">{req['Key']}</div>
                <div class="req-description">{req['Description']}</div>
            </div>
            """
    html_content += '</body></html>'
    with open(output_file, 'w') as f:
        f.write(html_content)


def execute_command(command_list):
    try:
        result = subprocess.run(
            command_list,
            capture_output=True,
            text=True,
            shell=False,  # More secure
        )
        if result.returncode != 0:
            logging.error(f'Error executing command: {" ".join(command_list)}')
            logging.error(result.stderr)
        else:
            logging.info(result.stdout)
        return result
    except Exception as e:
        logging.error(f'Failed to execute command: {e}')
        raise


def execute_rgw_commands(env_path, csv_path, export_repository):
    export_path = Path(export_repository)
    export_path.mkdir(parents=True, exist_ok=True)

    env_dir = Path(env_path).parent
    vectorcast_dir = os.environ.get('VECTORCAST_DIR', '')
    clicast = str(Path(vectorcast_dir) / 'clicast')
    if os.name == 'nt' and not clicast.endswith('.exe'):
        clicast += '.exe'

    # Convert paths to absolute and ensure proper formatting
    abs_export_path = str(export_path.resolve())
    abs_csv_path = str(Path(csv_path).resolve())

    rgw_prep_commands = [
        [clicast, '-lc', 'option', 'VCAST_REPOSITORY', abs_export_path],
        [clicast, '-lc', 'RGw', 'INitialize'],
        [clicast, '-lc', 'Rgw', 'Set', 'Gateway', 'CSV'],
        [clicast, '-lc', 'RGw', 'Configure', 'Set', 'CSV', 'csv_path', abs_csv_path],
        [clicast, '-lc', 'RGw', 'Configure', 'Set', 'CSV', 'use_attribute_filter', '0'],
        [clicast, '-lc', 'RGw', 'Configure', 'Set', 'CSV', 'filter_attribute'],
        [clicast, '-lc', 'RGw', 'Configure', 'Set', 'CSV', 'filter_attribute_value'],
        [clicast, '-lc', 'RGw', 'Configure', 'Set', 'CSV', 'id_attribute', 'ID'],
        [clicast, '-lc', 'RGw', 'Configure', 'Set', 'CSV', 'key_attribute', 'Key'],
        [clicast, '-lc', 'RGw', 'Configure', 'Set', 'CSV', 'title_attribute', 'Title'],
        [
            clicast,
            '-lc',
            'RGw',
            'Configure',
            'Set',
            'CSV',
            'description_attribute',
            'Description',
        ],
        [clicast, '-lc', 'RGw', 'Import'],
    ]

    # Change working directory before executing commands
    original_dir = os.getcwd()
    try:
        os.chdir(str(env_dir))
        for command in rgw_prep_commands:
            execute_command(command)
    finally:
        os.chdir(original_dir)


async def main(
    env_path,
    export_csv=None,
    export_html=None,
    export_repository=None,
    export_excel=None,
    json_events=False,
    combine_related_requirements=False,
    extended_reasoning=False,
    no_automatic_build=False,
    export_line_number=False,
    generate_high_level_requirements=False,
):
    log_level = os.environ.get('LOG_LEVEL', 'WARNING').upper()
    numeric_level = getattr(logging, log_level, logging.INFO)
    logging.basicConfig(level=numeric_level)

    if export_line_number:
        logging.warning(
            'Disabling post-processing of requirements to allow accurate export of covered line numbers'
        )

    environment = Environment(env_path, use_sandbox=False)

    if not environment.is_built:
        if no_automatic_build:
            logging.error(
                'Environment is not built and --no-automatic-build is set. Exiting.'
            )
            return
        else:
            logging.info('Environment is not built. Building it now...')
            environment.build()

    functions = environment.testable_functions

    generator = RequirementsGenerator(
        environment,
        combine_related_requirements=combine_related_requirements,
        extended_reasoning=extended_reasoning,
    )

    requirements = []

    # Initialize progress tracking
    total_functions = len(functions)
    processed_functions = 0

    async def generate_requirements(func):
        nonlocal processed_functions
        func_name = func['name']
        func_unit = func['unit_name']
        result, semantics = await generator.generate(
            func_name,
            post_process_requirements=not export_line_number,
            return_covered_semantic_parts=True,
        )
        processed_functions += 1
        progress = processed_functions / total_functions

        if json_events:
            print(json.dumps({'event': 'progress', 'value': progress}), flush=True)

        if result:
            for i, req in enumerate(result):
                req_id = f'{func_name}.{i + 1}'
                requirement = {
                    'Key': req_id,
                    'ID': req_id,
                    'Module': func_unit,
                    'Title': req,
                    'Description': req,
                    'Function': func_name,
                }
                if export_line_number:
                    requirement['Lines'] = repr(semantics[i].line_numbers)

                requirements.append(requirement)

    await tqdm_asyncio.gather(*[generate_requirements(func) for func in functions])

    if generate_high_level_requirements:
        high_level_generator = HighLevelRequirementsGenerator(
            environment,
            low_level_requirements=RequirementsManager(requirements),
            extended_reasoning=extended_reasoning,
        )

        async def generate_high_level_requirements(unit):
            unit_high_level_reqs = await high_level_generator.generate(unit)

            if unit_high_level_reqs:
                for i, req_text in enumerate(unit_high_level_reqs):
                    req_id = f'{unit}_HL.{i + 1}'  # Distinguish high-level req IDs
                    requirement = {
                        'Key': req_id,
                        'ID': req_id,
                        'Module': unit,
                        'Title': req_text,
                        'Description': req_text,
                        'Function': 'None',  # High-level requirements are not tied to a specific function
                    }
                    requirements.append(requirement)

        await tqdm_asyncio.gather(
            *[generate_high_level_requirements(unit) for unit in environment.units]
        )

    environment.cleanup()

    if export_csv:
        csv_path = export_csv
        save_requirements_to_csv(requirements, csv_path)
    else:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as temp_csv:
            csv_path = temp_csv.name
        save_requirements_to_csv(requirements, csv_path)

    if export_excel:
        save_requirements_to_excel(requirements, [environment], export_excel)

    if export_html:
        save_requirements_to_html(requirements, export_html)

    if export_repository:
        execute_rgw_commands(env_path, csv_path, export_repository)

    return generator.llm_client.total_cost['total_cost'], requirements


def cli():
    parser = argparse.ArgumentParser(
        description='Decompose design of functions into requirements.'
    )
    parser.add_argument(
        'env_path', help='Path to the VectorCAST environment directory.'
    )
    parser.add_argument(
        '--export-csv', help='Path to the output CSV file for requirements.'
    )
    parser.add_argument(
        '--export-html',
        help='Optional path to the output HTML file for pretty-printed requirements.',
    )
    parser.add_argument(
        '--export-excel',
        help='Optional path to the output Excel file for requirements.',
    )
    parser.add_argument(
        '--export-repository',
        help='Path to the VCAST_REPOSITORY for registering requirements.',
    )
    parser.add_argument(
        '--json-events', action='store_true', help='Output events in JSON format.'
    )
    parser.add_argument(
        '--overwrite-env',
        action='store_true',
        help='Prompt user for environment variables even if they are already set.',
    )
    parser.add_argument(
        '--combine-related-requirements',
        action='store_true',
        help='Combine related requirements into a single requirement after initial generation.',
    )
    parser.add_argument(
        '--extended-reasoning',
        action='store_true',
        help='Use extended reasoning for test generation.',
    )
    parser.add_argument(
        '--no-automatic-build',
        action='store_true',
        help='If the environment is not built, do not build it automatically.',
    )
    parser.add_argument(
        '--export-covered-lines',
        action='store_true',
        help=argparse.SUPPRESS,  # Controls if lines covered by the requirement are exported
    )
    parser.add_argument(
        '--generate-high-level-requirements',
        action='store_true',
        help='Also generate high-level requirements.',
    )

    args = parser.parse_args()

    asyncio.run(
        main(
            args.env_path,
            args.export_csv,
            args.export_html,
            args.export_repository,
            args.export_excel,
            json_events=args.json_events,
            extended_reasoning=args.extended_reasoning,
            no_automatic_build=args.no_automatic_build,
            export_line_number=args.export_covered_lines,
            generate_high_level_requirements=args.generate_high_level_requirements,
        )
    )


if __name__ == '__main__':
    cli()

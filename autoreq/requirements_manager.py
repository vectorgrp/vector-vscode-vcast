import logging
import re
import typing as t
import os
from pathlib import Path

import csv
import inspect
from openpyxl import load_workbook
import asyncio


class RequirementsManager:
    def __init__(self, requirements_or_path: t.Union[t.List[t.Dict], str]):
        # If it's a path (str), load from CSV or Excel; if it's already a list, store directly
        if isinstance(requirements_or_path, str) and os.path.isfile(
            requirements_or_path
        ):
            self._requirements = self._load_from_file(requirements_or_path)
        else:
            self._requirements = requirements_or_path  # Expecting a list of dicts

        filtered_requirements = []
        for req in self._requirements:
            if req['Function'] == 'None':
                logging.warning(
                    f'Requirement {req["ID"]} has no function assigned. Discarding.'
                )
                continue
            filtered_requirements.append(req)

        self._requirements = filtered_requirements

        # Build a quick lookup by requirement ID
        self._requirements_by_id = {}
        for req in self._requirements:
            self._requirements_by_id[req['ID']] = req

    def _load_from_csv(self, path: t.Union[str, Path]):
        reqs = []
        with open(path, newline='') as csvfile:
            reader = csv.DictReader(
                csvfile, skipinitialspace=True, quoting=csv.QUOTE_MINIMAL
            )
            for row in reader:
                reqs.append(row)
        return reqs

    def _load_from_excel(self, file_path: t.Union[str, Path]):
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        reqs = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            reqs.append(dict(zip(headers, row)))

        return reqs

    def _load_from_file(self, path: str):
        path_obj = Path(path)
        if path_obj.suffix == '.csv':
            return self._load_from_csv(path)
        elif path_obj.suffix == '.xlsx':
            return self._load_from_excel(path)
        else:
            raise ValueError(f'Unsupported file format: {path_obj.suffix}')

    @property
    def requirement_ids(self):
        return list(self._requirements_by_id.keys())

    def get_function(self, requirement_id):
        req = self._requirements_by_id.get(requirement_id)
        return req['Function'] if req else None

    def get_description(self, requirement_id):
        req = self._requirements_by_id.get(requirement_id)
        return req['Description'] if req else None

    def group_by_function(self, requirement_ids=None):
        if not requirement_ids:
            requirement_ids = self.requirement_ids

        grouped = {}
        for rid in requirement_ids:
            func = self.get_function(rid)
            if func not in grouped:
                grouped[func] = []
            grouped[func].append(rid)
        return grouped

    def get_requirements_for_function(self, function_name):
        return self.group_by_function().get(function_name, [])

    def requirements_to_dict(self):
        """Return a dictionary representation of all requirements."""
        return self._requirements

    def get_requirement(self, requirement_id):
        """Get a specific requirement by ID."""
        return self._requirements_by_id.get(requirement_id)

    def filter(self, filter_callback) -> 'RequirementsManager':
        """Filter requirements based on a callback function."""
        return RequirementsManager(
            [
                self._requirements_by_id[req_id]
                for req_id in filter(filter_callback, self.requirement_ids)
            ]
        )


class DecomposingRequirementsManager(RequirementsManager):
    def __init__(self, original_requirements, sub_requirements):
        """
        Initialize with original requirements and their decompositions.

        Args:
            original_requirements: A RequirementsManager instance with the original requirements
            sub_requirements: A dict mapping original requirement IDs to lists of sub-requirements
        """
        # Store the original requirements manager
        self.original_requirements = original_requirements

        # Store the sub-requirements mapping
        self.sub_requirements = sub_requirements

        # Create a flat list of all sub-requirements for RequirementsManager
        all_sub_requirements = []
        self._sub_to_original_map = {}  # Maps sub-requirement ID to original requirement ID

        for orig_id, sub_reqs in sub_requirements.items():
            for sub_req in sub_reqs:
                all_sub_requirements.append(sub_req)
                # Keep track of which original requirement this sub-requirement belongs to
                self._sub_to_original_map[sub_req['ID']] = orig_id

        # Initialize the parent class with the flattened sub-requirements
        super().__init__(all_sub_requirements)

    @staticmethod
    async def from_requirements_manager(
        original_requirements_manager, decomposer=lambda r: [r]
    ):
        """
        Create a DecomposingRequirementsManager from a RequirementsManager and a decomposer function.

        Args:
            original_requirements_manager: The original RequirementsManager
            decomposer: A function that takes a requirement and returns a list of sub-requirements

        Returns:
            A new DecomposingRequirementsManager instance
        """
        sub_requirements = {}
        is_async_decomposer = inspect.iscoroutinefunction(decomposer)

        async def process_requirement(req_id):
            original_req = original_requirements_manager.get_requirement(req_id)

            # Decompose the requirement
            if is_async_decomposer:
                decomposed_reqs = await decomposer(original_req)
            else:
                decomposed_reqs = decomposer(original_req)

            # Ensure each sub-requirement has a unique ID
            for i, sub_req in enumerate(decomposed_reqs):
                if 'ID' not in sub_req:
                    # If no ID is provided, create one based on the original ID
                    sub_req['ID'] = f'{original_req["ID"]}.{i + 1}'

            return req_id, decomposed_reqs

        # Create tasks for all requirements to process them in parallel
        tasks = [
            process_requirement(req_id)
            for req_id in original_requirements_manager.requirement_ids
        ]
        results = await asyncio.gather(*tasks)

        # Organize results into the sub_requirements dictionary
        for req_id, decomposed_reqs in results:
            sub_requirements[req_id] = decomposed_reqs

        return DecomposingRequirementsManager(
            original_requirements_manager, sub_requirements
        )

    @staticmethod
    async def from_file(original_requirements_path, decomposer=lambda r: [r]):
        """
        Create a DecomposingRequirementsManager from a CSV file and a decomposer function.

        Args:
            original_requirements_path: The path to the CSV file with original requirements
            decomposer: A function that takes a requirement and returns a list of sub-requirements

        Returns:
            A new DecomposingRequirementsManager instance
        """
        original_requirements = RequirementsManager(original_requirements_path)
        return await DecomposingRequirementsManager.from_requirements_manager(
            original_requirements, decomposer
        )

    def get_original_requirement(self, sub_requirement_id):
        """
        Get the original requirement that a sub-requirement belongs to.

        Args:
            sub_requirement_id: The ID of a sub-requirement

        Returns:
            The original requirement dictionary, or None if not found
        """
        if sub_requirement_id not in self._sub_to_original_map:
            return None

        orig_id = self._sub_to_original_map[sub_requirement_id]
        return self.original_requirements.get_requirement(orig_id)

    def get_original_requirement_id(self, sub_requirement_id):
        """
        Get the ID of the original requirement that a sub-requirement belongs to.

        Args:
            sub_requirement_id: The ID of a sub-requirement

        Returns:
            The ID of the original requirement, or None if not found
        """
        return self._sub_to_original_map.get(sub_requirement_id)

    def get_sub_requirements(self, original_requirement_id):
        """
        Get all sub-requirements for an original requirement.

        Args:
            original_requirement_id: The ID of an original requirement

        Returns:
            A list of sub-requirement dictionaries, or an empty list if not found
        """
        return self.sub_requirements.get(original_requirement_id, [])

    def get_sub_requirement_ids(self, original_requirement_id):
        """
        Get all sub-requirement IDs for an original requirement.

        Args:
            original_requirement_id: The ID of an original requirement

        Returns:
            A list of sub-requirement IDs, or an empty list if not found
        """
        sub_reqs = self.get_sub_requirements(original_requirement_id)
        return [req['ID'] for req in sub_reqs]

    def filter(self, filter_callback):
        """Ensure that we return a DecomposingRequirementsManager instance after filtering."""

        # Filter each of the sub-requirements
        filtered_sub_requirements = {}
        for orig_id, sub_reqs in self.sub_requirements.items():
            filtered_sub_requirements[orig_id] = [
                req for req in sub_reqs if filter_callback(req)
            ]

        return DecomposingRequirementsManager(
            self.original_requirements, filtered_sub_requirements
        )

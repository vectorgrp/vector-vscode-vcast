import json
import logging
from tempfile import NamedTemporaryFile, TemporaryDirectory
import typing as t
from pathlib import Path

import csv

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from pydantic import BaseModel, Field

from autoreq.util import (
    execute_command,
    get_vectorcast_cmd,
    temporary_working_directory,
)


class RequirementLocation(BaseModel):
    unit: t.Optional[str] = None
    function: t.Optional[str] = None
    lines: t.Optional[t.List[int]] = None


class Requirement(BaseModel):
    key: str
    id: str
    title: str
    description: str
    location: RequirementLocation = Field(default_factory=RequirementLocation)

    model_config = {
        'arbitrary_types_allowed': True,
        'ignored_types': (
            t.Callable,
            property,
        ),  # Tell Pydantic to ignore methods and properties
    }

    def to_flat_dict(self, convert_lines_to_string=False) -> dict:
        """Convert the requirement to a flat dictionary."""
        flat_dict = {
            'key': self.key,
            'id': self.id,
            'title': self.title,
            'description': self.description,
            'unit': self.location.unit or '',
            'function': self.location.function or '',
            'lines': self.location.lines,
        }

        if convert_lines_to_string:
            lines = self.location.lines
            flat_dict['lines'] = json.dumps(lines) if lines is not None else ''

        return flat_dict

    @staticmethod
    def from_flat_dict(data: dict):
        """Initialize the requirement from a flat dictionary."""

        assert all(key in data for key in ['key', 'id', 'title', 'description']), (
            "Data must contain 'key', 'id', 'title', and 'description' fields."
        )

        lines_str = data.get('lines')
        lines = None
        if lines_str and isinstance(lines_str, str):
            try:
                lines = json.loads(lines_str)
            except json.JSONDecodeError:
                logging.warning(
                    f"Failed to parse 'lines' field from string: {lines_str}"
                )
                lines = None
        elif isinstance(lines_str, list):
            lines = lines_str

        return Requirement(
            key=data['key'],
            id=data['id'],
            title=data['title'],
            description=data['description'],
            location=RequirementLocation(
                unit=data.get('unit') or None,
                function=data.get('function') or None,
                lines=lines,
            ),
        )


class DecomposedRequirement(Requirement):
    """
    A requirement that represents a sub-requirement of another requirement.
    It contains a reference to the original requirement it was decomposed from.
    """

    original_key: str

    def to_flat_dict(self, convert_lines_to_string=False) -> dict:
        data = super().to_flat_dict(convert_lines_to_string=convert_lines_to_string)
        data['original_key'] = self.original_key
        return data

    @staticmethod
    def from_flat_dict(data: dict):
        """Initialize the decomposed requirement from a flat dictionary."""
        req = Requirement.from_flat_dict(data)
        return DecomposedRequirement(
            key=req.key,
            id=req.id,
            title=req.title,
            description=req.description,
            location=req.location,
            original_key=data['original_key'],
        )


class RequirementsCollection:
    def __init__(self, requirements=None):
        self.requirements = requirements or []
        self._validate_requirements()

    @property
    def _key_requirement_map(self):
        return {req.key: req for req in self.requirements}

    def _validate_requirements(self):
        """
        Validate that all requirements have unique keys.
        """
        if not all(isinstance(req, Requirement) for req in self.requirements):
            raise TypeError('All items in requirements must be of type Requirement.')

        if len(self.requirements) != len(self._key_requirement_map):
            raise ValueError('Requirements must have unique keys.')

    @property
    def requirement_keys(self):
        return [req.key for req in self.requirements]

    def __getitem__(self, item):
        matching_req = self._key_requirement_map.get(item)

        if matching_req is None:
            raise KeyError(f"Requirement with ID '{item}' not found.")
        else:
            return matching_req

    def __iter__(self):
        return iter(self.requirements)

    def __len__(self):
        return len(self.requirements)

    def __contains__(self, x):
        return x in self.requirements

    def filter(self, filter_callback):
        return RequirementsCollection(
            [req for req in self.requirements if filter_callback(req)]
        )

    def group_by(self, group_callback):
        grouped = {}
        for req in self.requirements:
            key = group_callback(req)
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(req)
        return grouped

    def map(self, map_callback):
        return RequirementsCollection([map_callback(req) for req in self.requirements])

    def flat_map(self, map_callback):
        """
        Apply a mapping function to each requirement and flatten the results.
        """
        return RequirementsCollection(
            [item for req in self.requirements for item in map_callback(req)]
        )

    def append(self, requirement):
        """
        Append a new requirement to the collection.
        """
        self.requirements.append(requirement)
        self._validate_requirements()

    def extend(self, requirements):
        """
        Extend the collection with a list of new requirements.
        """
        if isinstance(requirements, RequirementsCollection):
            self.requirements.extend(requirements.requirements)
        elif isinstance(requirements, list):
            self.requirements.extend(requirements)
        else:
            raise TypeError(
                'Requirements must be a RequirementsCollection or a list of Requirement objects.'
            )
        self._validate_requirements()

    @staticmethod
    def from_csv(path):
        raw_reqs = []
        with open(path, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(
                csvfile, skipinitialspace=True, quoting=csv.QUOTE_MINIMAL
            )
            for row in reader:
                raw_reqs.append(
                    {
                        key.lower() if key != 'Module' else 'unit': value
                        for key, value in row.items()
                    }
                )
                # TODO: Get rid of the legacy patching code when possible

        reqs = [Requirement.from_flat_dict(req) for req in raw_reqs]
        return RequirementsCollection(reqs)

    def to_csv(self, path):
        with open(path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'key',
                'id',
                'title',
                'description',
                'unit',
                'function',
                'lines',
            ]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for req in self.requirements:
                writer.writerow(req.to_flat_dict(convert_lines_to_string=True))

    def to_html(self, output_file=None):
        html_content = """
        <html>
        <head>
            <title>Requirements</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; background-color: #ffffff; color: #000000; }
                h1 { color: #2c3e50; }
                h2 { color: #34495e; margin-top: 30px; }
                .requirement { background-color: #f7f7f7; padding: 15px; margin: 10px 0; border-radius: 5px; }
                .req-key { font-weight: bold; color: #2980b9; }
                .req-description { margin-top: 10px; color: #333333; }
            </style>
        </head>
        <body>
            <h1>Requirements</h1>
        """
        requirements_by_function = self.group_by(
            lambda req: req.location.function or 'No Function'
        )
        for function, reqs in requirements_by_function.items():
            html_content += f'<h2>{function}</h2>'
            for req in reqs:
                html_content += f"""
                <div class="requirement">
                    <div class="req-key">{req.key}</div>
                    <div class="req-description">{req.description}</div>
                </div>
                """

        html_content += '</body></html>'

        if output_file:
            with open(output_file, 'w') as f:
                f.write(html_content)
        else:
            return html_content

    def _fix_worksheet_cols_width(self, ws):
        """Auto-adjust column widths based on content."""
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max(
                        (dims.get(cell.column, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value

    def to_excel(
        self, output_path: str, include_metadata: bool = True, source_envs=None
    ) -> None:
        """Export requirements to an Excel file with dropdowns and formatting."""
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = 'Requirements'

        # Get function data from source environments if provided
        funcs = {}
        if source_envs:
            for env in source_envs:
                for f in env.testable_functions:
                    key = f['unit_name']
                    funcs.setdefault(key, [])
                    funcs[key].append(f['name'])

        # Define headers with metadata
        headers = ['Key', 'ID', 'Title', 'Description']
        if include_metadata:
            headers.extend(['Module', 'Function', 'Lines'])

        # Create header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(
            fill_type='solid', start_color='C6EFCE'
        )  # Light green

        # Write headers with styling
        for col, header in enumerate(headers, 1):
            if ws is not None:
                cell = ws.cell(row=1, column=col)
                if cell is not None:
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill

        # Write requirement data
        for row_idx, req in enumerate(self.requirements, 2):
            if ws is not None:
                ws.cell(row=row_idx, column=1).value = req.key
                ws.cell(row=row_idx, column=2).value = req.id
                ws.cell(row=row_idx, column=3).value = req.title
                ws.cell(row=row_idx, column=4).value = req.description

                if include_metadata:
                    ws.cell(row=row_idx, column=5).value = (
                        req.location.unit if req.location else ''
                    )
                    ws.cell(row=row_idx, column=6).value = (
                        req.location.function if req.location else ''
                    )
                    ws.cell(row=row_idx, column=7).value = (
                        ', '.join(map(str, req.location.lines))
                        if req.location and req.location.lines
                        else ''
                    )

        # Add dropdowns if source_envs provided and metadata included
        if source_envs and include_metadata and funcs and ws is not None:
            # Create hidden sheet for dropdown options
            lists_ws = wb.create_sheet('Options')
            lists_ws.sheet_state = 'hidden'

            # List all modules in column A
            modules = list(funcs.keys())
            for row, module in enumerate(modules, start=1):
                lists_ws.cell(row=row, column=1, value=module)

            # For each module, write its functions in its own column and create named range
            for i, module in enumerate(modules):
                functions = funcs[module] + ['None']
                col = i + 2
                for row, func in enumerate(functions, start=1):
                    lists_ws.cell(row=row, column=col, value=func)
                col_letter = get_column_letter(col)
                end_row = len(functions)
                range_ref = f'${col_letter}$1:${col_letter}${end_row}'
                range_name = module.replace(' ', '_').replace('-', '_')
                wb.create_named_range(range_name, lists_ws, range_ref)

            # Add data validation for Unit column (assuming it's column 5)
            num_data_rows = len(self.requirements)
            unit_col_index = 5  # Unit column
            function_col_index = 6  # Function column
            unit_col_letter = get_column_letter(unit_col_index)

            # Module/Unit dropdown
            dv_unit = DataValidation(
                type='list',
                formula1=f'=Options!$A$1:$A${len(modules)}',
                allow_blank=True,
            )
            ws.add_data_validation(dv_unit)
            dv_unit_range = f'{get_column_letter(unit_col_index)}2:{get_column_letter(unit_col_index)}{num_data_rows + 1}'
            dv_unit.add(dv_unit_range)

            # Function dropdown using INDIRECT
            for row in range(2, num_data_rows + 2):
                unit_cell = f'{unit_col_letter}{row}'
                formula = f'=INDIRECT(SUBSTITUTE({unit_cell}," ","_"))'
                dv_func = DataValidation(
                    type='list', formula1=formula, allow_blank=True
                )
                ws.add_data_validation(dv_func)
                function_cell = f'{get_column_letter(function_col_index)}{row}'
                dv_func.add(function_cell)

        # Auto-adjust column widths
        if ws is not None:
            self._fix_worksheet_cols_width(ws)

        # Save the workbook
        wb.save(output_path)

    @staticmethod
    def from_excel(excel_path: str):
        """Load requirements from an Excel file."""
        wb = load_workbook(excel_path)
        ws = wb.active

        # Get headers from first row
        headers = []
        if ws is not None:
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

        # Clear current requirements
        requirements = []

        # Read data rows
        if ws is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue

                # Create requirement data dict
                req_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and value is not None:
                        req_data[headers[i]] = str(value).strip() if value else ''

                # Extract basic requirement info
                key = req_data.get('Key', '')
                req_id = req_data.get('ID', '')
                title = req_data.get('Title', '')
                description = req_data.get('Description', '')

                if not key or not description:
                    continue

                # Parse location info
                location = None
                unit = req_data.get('Module', '')
                function = req_data.get('Function', '')
                lines_str = req_data.get('Lines', '')

                if function == 'None':
                    function = None

                if unit or function or lines_str:
                    lines = []
                    if lines_str:
                        try:
                            lines = [
                                int(line.strip())
                                for line in lines_str.split(',')
                                if line.strip().isdigit()
                            ]
                        except ValueError:
                            lines = []

                    location = RequirementLocation(
                        unit=unit if unit else None,
                        function=function if function else None,
                        lines=lines if lines else None,
                    )

                # Create requirement
                requirement = Requirement(
                    key=key,
                    id=req_id,
                    title=title,
                    description=description,
                    location=location or RequirementLocation(),
                )

                requirements.append(requirement)

        return RequirementsCollection(requirements)

    @classmethod
    def from_path(cls, path: str) -> 'RequirementsCollection':
        """Load requirements from a file path, auto-detecting format."""
        path_obj = Path(path)

        if path_obj.is_dir():
            return RequirementsCollection.from_rgw(path)
        elif path_obj.suffix.lower() == '.xlsx':
            return RequirementsCollection.from_excel(path)
        elif path_obj.suffix.lower() == '.csv':
            return cls.from_csv(path)
        else:
            raise ValueError(f'Unsupported file format: {path_obj.suffix}')

    def to_rgw(self, rgw_path, target_env=None, only_traceability=False):
        # Get traceability path first
        requirements_traceability_path = self._get_rgw_requirements_traceability_path(
            rgw_path, check_exist=False
        )

        if not only_traceability:
            with TemporaryDirectory() as temp_dir:
                execution_path = target_env.env_dir if target_env else temp_dir
                with temporary_working_directory(execution_path):
                    self._create_rgw(rgw_path)

        elif requirements_traceability_path:
            Path(requirements_traceability_path).parent.mkdir(
                parents=True, exist_ok=True
            )

        # Now save traceability info
        if requirements_traceability_path:
            traceability_data = {}
            for req in self.requirements:
                traceability_data[req.key] = {
                    'unit': req.location.unit,
                    'function': req.location.function,
                    'lines': req.location.lines,
                }

            with open(str(requirements_traceability_path), 'w') as f:
                json.dump(traceability_data, f, indent=4)

    def _create_rgw(self, rgw_path):
        rgw_path = Path(rgw_path)
        rgw_path.mkdir(parents=True, exist_ok=True)

        with NamedTemporaryFile(mode='w', suffix='.csv') as temp_file:
            self.to_csv(temp_file.name)

            rgw_prep_commands = [
                ['-lc', 'option', 'VCAST_REPOSITORY', str(rgw_path.resolve())],
                ['-lc', 'RGw', 'INitialize'],
                ['-lc', 'Rgw', 'Set', 'Gateway', 'CSV'],
                [
                    '-lc',
                    'RGw',
                    'Configure',
                    'Set',
                    'CSV',
                    'csv_path',
                    temp_file.name,
                ],
                [
                    '-lc',
                    'RGw',
                    'Configure',
                    'Set',
                    'CSV',
                    'use_attribute_filter',
                    '0',
                ],
                ['-lc', 'RGw', 'Configure', 'Set', 'CSV', 'filter_attribute'],
                ['-lc', 'RGw', 'Configure', 'Set', 'CSV', 'filter_attribute_value'],
                ['-lc', 'RGw', 'Configure', 'Set', 'CSV', 'id_attribute', 'id'],
                ['-lc', 'RGw', 'Configure', 'Set', 'CSV', 'key_attribute', 'key'],
                [
                    '-lc',
                    'RGw',
                    'Configure',
                    'Set',
                    'CSV',
                    'title_attribute',
                    'title',
                ],
                [
                    '-lc',
                    'RGw',
                    'Configure',
                    'Set',
                    'CSV',
                    'description_attribute',
                    'description',
                ],
                ['-lc', 'RGw', 'Import'],
            ]

            rgw_prep_commands = list(
                map(
                    lambda args: get_vectorcast_cmd('clicast', args),
                    rgw_prep_commands,
                )
            )

            for cmd in rgw_prep_commands:
                execute_command(cmd)

    @staticmethod
    def from_rgw(rgw_path):
        requirements_json_path = RequirementsCollection._get_rgw_requirements_json_path(
            rgw_path
        )

        assert requirements_json_path is not None, (
            'The requirements gateway does not contain a file containing requirements.'
        )
        with open(requirements_json_path, 'r') as f:
            data = json.load(f)

        # For older versions of VectorCAST, the requirements might be under 'requirements' key.
        if 'requirements' in data:
            data = data['requirements']

        req_dicts = []
        for group_id, reqs_info in data.items():
            for req_key, req in reqs_info.items():
                req['key'] = req_key
                req_dicts.append(req)

        requirements_traceability_path = (
            RequirementsCollection._get_rgw_requirements_traceability_path(rgw_path)
        )

        if requirements_traceability_path is not None:
            with open(requirements_traceability_path, 'r') as f:
                traceability_data = json.load(f)

            for req in req_dicts:
                traceability_info = traceability_data.get(req['key'])

                if traceability_data is None:
                    logging.warning(
                        f"Traceability info for requirement '{req['key']}' not found in traceability data."
                    )
                    continue

                for traceability_key, traceability_value in traceability_info.items():
                    req[traceability_key] = traceability_value

        reqs = [Requirement.from_flat_dict(req) for req in req_dicts]
        return RequirementsCollection(reqs)

    @staticmethod
    def _get_rgw_requirements_json_path(rgw_path, check_exist=True):
        rgw_path = Path(rgw_path)

        requirements_json_path = rgw_path / 'requirements_gateway' / 'requirements.json'
        if not check_exist or requirements_json_path.is_file():
            return requirements_json_path

        requirements_json_path = rgw_path / 'requirements_gateway' / 'repository.json'
        if not check_exist or requirements_json_path.is_file():
            logging.warning(
                'The requirements gateway contains a repository.json file instead of requirements.json.\n'
                'In newer versions of VectorCAST, the requirements gateway uses requirements.json instead of repository.json.'
            )
            return requirements_json_path

    @staticmethod
    def _get_rgw_requirements_traceability_path(rgw_path, check_exist=True):
        rgw_path = Path(rgw_path)

        requirements_traceability_path = (
            rgw_path / 'requirements_gateway' / 'traceability.json'
        )
        if not check_exist or requirements_traceability_path.is_file():
            return requirements_traceability_path

        return None
